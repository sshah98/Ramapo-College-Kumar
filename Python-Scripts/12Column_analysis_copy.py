import pandas as pd
import numpy as np
from openpyxl import load_workbook


fileName = "SelectionSelfExpln" #copy/paste filename without file extension
file = "%s.xlsx" % (fileName) #creates file variable for rest of program

path = r'/home/suraj/Documents/Ramapo-Project/SelectionSelfExpln.xlsx' #path to copy loaded workbook to

correctScore = 1.0 #correct score for each question. Can be changed

def pretest(): #function that creates new sheet with only pretest data
    print("Reading data...")

    data_df = pd.read_excel(file, sheetname='Data') #creates new dataframe to read first sheet
    data_df = data_df[data_df['Stage'].str.contains('Pretest', na=False)] #checks for only pretest data
    print("Gathering all pretest rows")
    print("Saving data to pretest sheet")
    writer = pd.ExcelWriter(path, engine='openpyxl') #in all functions - writes to a new sheet using the dataframe created in function
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data_df.to_excel(writer, sheet_name='Pretest')
    writer.save()

def grades():
    print("Reading data from pretest")

    pretest_df = pd.read_excel(file, sheetname='Pretest') #creates new dataframe using Pretest sheet
    print("keeping only the columns necessary")

    pretest_df = pretest_df.iloc[:, [10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54]] #locates only the score columns to copy those

    pretest_df = pretest_df.drop(pretest_df.columns[12:], axis=1) #drops all other extra columns after the last score
    pretest_df.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
    print("Saving data to grades sheet")

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    pretest_df.to_excel(writer, sheet_name='Grades')
    writer.save()

def allgrades():
    grades_df = pd.read_excel(file, sheetname='Grades')
    grades_df = grades_df[grades_df.columns[0:12]].replace('', np.nan)
    grades_df = grades_df.dropna()
    print("Opening necessary dataframes for AllGrades")

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    grades_df.to_excel(writer, sheet_name='AllGrades')
    writer.save()

def correct():
    allGrades_df = pd.read_excel(file, sheetname='AllGrades')

    allGrades_df[allGrades_df != 1] = 0
    allGrades_df = allGrades_df.replace(0, '')
    print("Replaced all empty cells with 0 and saving to new sheet named Correct")

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    allGrades_df.to_excel(writer, sheet_name='Correct')
    writer.save()

def conditional():
    correct_df = pd.read_excel(file, sheetname='Correct')

    for i in range(11, 0, -1):
        correct_df[correct_df.columns[i]] = np.where((correct_df[correct_df.columns[i-1]] == correctScore) & (correct_df[correct_df.columns[i]] == correctScore), 1, '')
    print("Finding all correct rows and conditionally replacing incorrect cells with empty string")

    correct_df['1'] = correct_df['1'].replace([1, np.nan], '')
    correct_df = correct_df.apply(pd.to_numeric)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    correct_df.to_excel(writer, sheet_name='Conditional')
    writer.save()

def cumulative():
    df = pd.read_excel(file, sheetname='Correct')
    df1 = pd.read_excel(file, sheetname='Conditional')
    df2 = pd.DataFrame(columns=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'])
    df2 = df2.fillna(0)
    print("using dataframes to solve for cumulative sheet")

    df2['1'] = df['1']
    df2['2'] = df1['2']

    df2 = df2.fillna('')
    df2 = df2.apply(pd.to_numeric)

    for i in range(2, 12, 1):
        df2[df2.columns[i]] = np.where((df2[df2.columns[i-1]] == correctScore) & (df1[df1.columns[i]] == correctScore), 1, '')
        df2 = df2.apply(pd.to_numeric)

    df2 = df2.apply(pd.to_numeric)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df2.to_excel(writer, sheet_name='Cumulative')
    writer.save()

def stats():
    print("Accessing data from sheets to store in dataframe")

    correct_df = pd.read_excel(file, sheetname='Correct')
    allgrades_df = pd.read_excel(file, sheetname='AllGrades')
    conditional_df = pd.read_excel(file, sheetname='Conditional')
    cumulative_df = pd.read_excel(file, sheetname='Cumulative')

    allGradesTot = allgrades_df[allgrades_df.columns[0:12]].count()
    correctGradesSum = correct_df[correct_df.columns[0:12]].sum()

    subset2 = []
    for i in range(11):
        conditional_val = conditional_df[conditional_df.columns[i + 1]].sum()
        correct_val = correct_df[correct_df.columns[i]].sum()
        i += 1
        subset2.append(conditional_val / correct_val)

    conditional_df_1 = pd.DataFrame(subset2)
    pd.melt(conditional_df_1)
    conditional_df_1 = conditional_df_1.T
    conditional_df_1.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

    subset3 = []
    correct_sum = correct_df[correct_df.columns[0:12]].sum()
    conditional_sum = conditional_df[conditional_df.columns[0:12]].sum()
    subset3.append(conditional_sum / correct_sum)

    conditional_df2 = pd.DataFrame(subset3,columns=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'])
    conditional_df2 = conditional_df2.shift(periods=-1, axis=1)

    result = pd.concat([conditional_df_1, conditional_df2], axis=0)
    result = result[['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']]

    numAnswered = allgrades_df[allgrades_df.columns[0]].count()

    subset5 = []
    for i in range(11):
        subset5.append(cumulative_df[cumulative_df.columns[i + 1]].sum() / cumulative_df[cumulative_df.columns[i]].sum())
        i += 1

    cuml_df2 = pd.DataFrame(subset5)
    pd.melt(cuml_df2)
    cuml_df2 = cuml_df2.T
    cuml_df2.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

    probA_Incorrect = []
    for i in range(11):
        a = float(sum(pd.isnull(correct_df[correct_df.columns[i]]) & pd.notnull(correct_df[correct_df.columns[i+1]])))
        b = float(sum(pd.isnull(correct_df[correct_df.columns[i]])))
        probA_Incorrect.append(a/b)

    probABCorrect_df = pd.DataFrame(probA_Incorrect)
    pd.melt(probABCorrect_df)
    probABCorrect_df = probABCorrect_df.T
    probABCorrect_df.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
        
    
    result = pd.concat([conditional_df_1, conditional_df2, cuml_df2, probABCorrect_df], axis=0)
    result = result[['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']]
    
    tot = "N: " + str(numAnswered)
    result.iloc[0,0] = tot
    
    print(result)
    print("Saving data to Stats sheet")

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    result.to_excel(writer, sheet_name='Stats')
    writer.save()

def legend():

    info = {
    'A. Follows order from previous sheet': 'Values from top to bottom on previous sheet',
    'B. P(n)': 'Probability of the problem being solved correctly',
    'C. P(n|n - 1)': 'Conditional probability of solving (n+1)th problem correctly given nth problem was solved correctly',
    'D. P(n-1|n)': 'Calculated by dividing P(n-1) / P(correct)',
    'E. P(n,n-1,n-2,...,1)' : 'cumulative probability 1',
    'F. P(n|n-1,n-2,...,1)': 'cumulative probability 2',
    'G. P(B| ~A)': 'Probability of B correct given A incorrect',
    'H. P(n|n - 1) * P(n - 1)': 'Verification of data 1',
    'I. P(n - 1|n) * P(n)': 'Verification of data 2',
    'J. P(n|n-1,n-2,...,1)': 'Verification of chain rule'
    }

    df = pd.DataFrame.from_records(info, index=[0]).T

    df = df.sort_index(ascending=True)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name='Legend')
    writer.save()


pretest()
grades()
allgrades()
correct()
conditional()
cumulative()
stats()
legend()