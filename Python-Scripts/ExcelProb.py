import pandas as pd
import numpy as np
from openpyxl import load_workbook
import docx

fileName = "ArithmeticS14VizExplnANONUploaded" #copy/paste filename without file extension
file = "%s.xlsx" % (fileName) #creates file variable for rest of program

path = r'/home/suraj/Documents/Ramapo-Project/ArithmeticS14VizExplnANONUploaded.xlsx' #path to copy loaded workbook to

correctScore = 1.0 #correct score for each question. Can be changed

def pretest(): #function that creates new sheet with only pretest data
    
    print "Reading data..."
    data_df = pd.read_excel(file, sheetname='Data') #creates new dataframe to read first sheet
    data_df = data_df[data_df['Stage'].str.contains('Pretest', na=False)] #checks for only pretest data

    print "Writing to pretest data..."
    writer = pd.ExcelWriter(path, engine='openpyxl') #in all functions - writes to a new sheet using the dataframe created in function
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data_df.to_excel(writer, sheet_name='Pretest')
    writer.save()

def grades():

    pretest_df = pd.read_excel(file, sheetname='Pretest') #creates new dataframe using Pretest sheet

    pretest_df = pretest_df.iloc[:, [10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54, 58, 62, 66, 70]] #locates only the score columns to copy those

    pretest_df = pretest_df.drop(pretest_df.columns[16:], axis=1) #drops all other extra columns after the last score
    pretest_df.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    pretest_df.to_excel(writer, sheet_name='Grades')
    writer.save()

def allgrades():
    grades_df = pd.read_excel(file, sheetname='Grades')
    grades_df = grades_df[grades_df.columns[0:16]].replace('', np.nan)
    grades_df = grades_df.dropna()

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    grades_df.to_excel(writer, sheet_name='AllGrades')
    writer.save()

def correct():
    allGrades_df = pd.read_excel(file, sheetname='AllGrades')

    allGrades_df[allGrades_df != 1] = 0
    allGrades_df = allGrades_df.replace(0, '')

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    allGrades_df.to_excel(writer, sheet_name='Correct')
    writer.save()

def conditional():
    correct_df = pd.read_excel(file, sheetname='Correct')

    for i in xrange(15, 0, -1):
        correct_df[correct_df.columns[i]] = np.where((correct_df[correct_df.columns[i-1]] == correctScore) & (correct_df[correct_df.columns[i]] == correctScore), 1, '')

    correct_df['1'] = correct_df['1'].replace([1, np.nan], '')
    correct_df = correct_df.apply(pd.to_numeric)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    correct_df.to_excel(writer, sheet_name='Conditional')
    writer.save()

def cumulative():
    df = pd.read_excel(file, sheetname='Correct')
    df1 = pd.read_excel(file, sheetname='Conditional')
    df2 = pd.DataFrame(columns=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16'])
    df2 = df2.fillna(0)

    df2['1'] = df['1']
    df2['2'] = df1['2']

    df2 = df2.fillna('')
    df2 = df2.apply(pd.to_numeric)

    for i in xrange(2, 16, 1):
        df2[df2.columns[i]] = np.where((df2[df2.columns[i-1]] == correctScore) & (df1[df1.columns[i]] == correctScore), 1, '')
        df2 = df2.apply(pd.to_numeric)

    df2 = df2.apply(pd.to_numeric)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df2.to_excel(writer, sheet_name='Cumulative')
    writer.save()

def stats():

    correct_df = pd.read_excel(file, sheetname='Correct')
    allgrades_df = pd.read_excel(file, sheetname='AllGrades')
    conditional_df = pd.read_excel(file, sheetname='Conditional')
    cumulative_df = pd.read_excel(file, sheetname='Cumulative')

    allGradesTot = allgrades_df[allgrades_df.columns[0:16]].count()
    correctGradesSum = correct_df[correct_df.columns[0:16]].sum()

    subset1 = []
    subset1.append(correctGradesSum / allGradesTot)
    probability_df = pd.DataFrame(subset1,columns=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16'])

    subset2 = []
    for i in xrange(15):
        conditional_val = conditional_df[conditional_df.columns[i + 1]].sum()
        correct_val = correct_df[correct_df.columns[i]].sum()
        i += 1
        subset2.append(conditional_val / correct_val)

    conditional_df_1 = pd.DataFrame(subset2)
    pd.melt(conditional_df_1)
    conditional_df_1 = conditional_df_1.T
    conditional_df_1.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']

    subset3 = []
    correct_sum = correct_df[correct_df.columns[0:16]].sum()
    conditional_sum = conditional_df[conditional_df.columns[0:16]].sum()
    subset3.append(conditional_sum / correct_sum)

    conditional_df2 = pd.DataFrame(subset3,columns=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15','16'])
    conditional_df2 = conditional_df2.shift(periods=-1, axis=1)

    result = pd.concat([probability_df, conditional_df_1, conditional_df2], axis=0)
    result = result[['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']]

    numAnswered = allgrades_df[allgrades_df.columns[0]].count()

    subset4 = []
    for i in xrange(16):
        subset4.append((cumulative_df[cumulative_df.columns[i]].sum()) / numAnswered)
        i += 1

    cuml_df_1 = pd.DataFrame(subset4)
    pd.melt(cuml_df_1)
    cuml_df_1 = cuml_df_1.T
    cuml_df_1.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']

    subset5 = []
    for i in xrange(15):
        subset5.append(cumulative_df[cumulative_df.columns[i + 1]].sum() / cumulative_df[cumulative_df.columns[i]].sum())
        i += 1

    cuml_df2 = pd.DataFrame(subset5)
    pd.melt(cuml_df2)
    cuml_df2 = cuml_df2.T
    cuml_df2.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']

    subset6 = []
    for i in xrange(15):
        subset6.append(result.iloc[1, i + 1] * result.iloc[0, i])
        i += 1
    verification_df1 = pd.DataFrame(subset6)
    pd.melt(verification_df1)
    verification_df1 = verification_df1.T
    verification_df1.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15']

    subset7 = []
    for i in xrange(15):
        subset7.append(result.iloc[2, i] * result.iloc[0, i + 1])
        i += 1
    verification_df_2 = pd.DataFrame(subset7)
    pd.melt(verification_df_2)
    verification_df_2 = verification_df_2.T
    verification_df_2.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15']

    result = pd.concat([probability_df, conditional_df_1, conditional_df2, cuml_df_1, cuml_df2, verification_df1, verification_df_2], axis=0)
    result = result[['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']]

    chainruleSubset = []
    chainruleSubset.append(result.iloc[3, 1] / result.iloc[0, 0])
    chainruleSubset.append(result.iloc[3, 2] / (result.iloc[1, 1] * result.iloc[0, 0]))
    chainruleSubset.append((result.iloc[3, 3]) / (result.iloc[4, 2] * result.iloc[1, 1] * result.iloc[0, 0]))

    for i in xrange(1, 13, 1):
        chainruleSubset.append((result.iloc[3, i+3]) / (np.prod((result.iloc[4,1:i+3])) * result.iloc[0,0]))

    chainrule_verif_df = pd.DataFrame(chainruleSubset)
    pd.melt(chainrule_verif_df)
    chainrule_verif_df = chainrule_verif_df.T
    chainrule_verif_df.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']


    probA_Incorrect = []
    for i in xrange(15):
        a = float(sum(pd.isnull(correct_df[correct_df.columns[i]]) & pd.notnull(correct_df[correct_df.columns[i+1]])))
        b = float(sum(pd.isnull(correct_df[correct_df.columns[i]])))
        probA_Incorrect.append(a/b)

    probABCorrect_df = pd.DataFrame(probA_Incorrect)
    pd.melt(probABCorrect_df)
    probABCorrect_df = probABCorrect_df.T
    probABCorrect_df.columns = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']

    stringDataframe = pd.DataFrame(['1,2', '2,3', '3,4', '4,5', '5,6', '6,7', '7,8', '8,9', '9,10', '10,11', '11, 12', '12,13', '13,14', '14,15', '15,16'])
    pd.melt(stringDataframe)
    stringDataframe = stringDataframe.T
    stringDataframe.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15']

    result = pd.concat([probability_df, conditional_df_1, conditional_df2, cuml_df_1, cuml_df2, probABCorrect_df, stringDataframe, verification_df1, verification_df_2, chainrule_verif_df], axis=0)
    result = result[['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16']]

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    result.to_excel(writer, sheet_name='Stats')
    writer.save()


def legend():

    info = {
    'A. Follows order from previous sheet': 'Values from top to bottom on previous sheet',
    'B. P(n)': 'Probability of the problem being solved correctly',
    'C. P(n|n - 1)': 'Conditional probability of solving (n+1)th problem correctly given nth problem was solved correctly',
    'D. P(n-1|n)': 'Calculated by dividing P(n-1) / P(correct)',
    'E. P(n,n-1,n-2,...,1)' : 'cumulative probability 1',
    'F. P(n|n-1,n-2,...,1)': 'cumulative probability 2',
    'G. P(B| ~A)': 'Probability of B correct given A incorrect',
    'H. P(n|n - 1) * P(n - 1)': 'Verification of data 1',
    'I. P(n - 1|n) * P(n)': 'Verification of data 2',
    'J. P(n|n-1,n-2,...,1)': 'Verification of chain rule'
    }

    df = pd.DataFrame.from_records(info, index=[0]).T

    df = df.sort_index(ascending=True)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    book = load_workbook(path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name='Legend')
    writer.save()




# pretest()
# grades()
# allgrades()
# correct()
# conditional()
# cumulative()
# stats()
# legend()

